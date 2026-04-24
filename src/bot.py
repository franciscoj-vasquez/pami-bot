from playwright.sync_api import Playwright, sync_playwright, TimeoutError as PWTimeout
from dotenv import load_dotenv
from getpass import getpass
from datetime import datetime, date
from pathlib import Path
import pandas as pd
import random
import time
import re
import os
import winreg
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def _get_documents_dir() -> Path:
    with winreg.OpenKey(
        winreg.HKEY_CURRENT_USER,
        r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders",
    ) as key:
        return Path(winreg.QueryValueEx(key, "Personal")[0])


load_dotenv()

USUARIO    = os.getenv("PAMI_USER") or input("Usuario PAMI: ")
CLAVE      = os.getenv("PAMI_PASS") or getpass("Contraseña PAMI: ")
DRY_RUN    = os.getenv("PAMI_DRY_RUN",  "") not in ("", "0", "false")
HEADLESS   = os.getenv("PAMI_HEADLESS", "1") not in ("", "0", "false")
PAMI_DIR       = _get_documents_dir() / "Ordenes PAMI"
PAMI_PACIENTES = PAMI_DIR / "pacientes"
PAMI_REPORTES  = PAMI_DIR / "reportes"
EXCEL_PATH     = Path(os.getenv("PAMI_EXCEL")) if os.getenv("PAMI_EXCEL") else PAMI_PACIENTES / "pacientes.xlsx"
STOP_FLAG      = Path("stop.flag")  # relativo al cwd = data/
RETRIES        = int(os.getenv("PAMI_RETRIES", "0"))

_PROFILES = {
    "cauteloso": dict(slow_mo=250, pausa_min=0.8,  pausa_max=2.0,  corta_min=0.4, corta_max=0.8,  typing=120),
    "normal":    dict(slow_mo=150, pausa_min=0.4,  pausa_max=1.0,  corta_min=0.2, corta_max=0.5,  typing=80),
    "rapido":    dict(slow_mo=80,  pausa_min=0.15, pausa_max=0.4,  corta_min=0.1, corta_max=0.25, typing=50),
}
_p        = _PROFILES.get(os.getenv("PAMI_SPEED", "normal"), _PROFILES["normal"])
SLOW_MO   = _p["slow_mo"]
PAUSA_MIN = _p["pausa_min"]
PAUSA_MAX = _p["pausa_max"]
CORTA_MIN = _p["corta_min"]
CORTA_MAX = _p["corta_max"]
TYPING_MS = _p["typing"]

class LoginError(Exception):
    pass

class OrdenError(Exception):
    """Error recuperable: se cancela la orden y se continúa con la siguiente."""
    pass

class BeneficioNoEncontrado(OrdenError):
    pass

class DiagnosticoNoEncontrado(OrdenError):
    pass

class PracticaNoEncontrada(OrdenError):
    pass

class SesionDuplicada(OrdenError):
    pass

class DetenerError(Exception):
    """Señal de parada solicitada por el usuario — cancela la orden y genera el reporte."""
    pass

def pausa(minimo=None, maximo=None):
    time.sleep(random.uniform(
        minimo if minimo is not None else PAUSA_MIN,
        maximo if maximo is not None else PAUSA_MAX,
    ))

def pausa_corta():
    time.sleep(random.uniform(CORTA_MIN, CORTA_MAX))

def check_stop():
    if STOP_FLAG.exists():
        raise DetenerError()

def leer_pacientes():
    df = pd.read_excel(EXCEL_PATH, dtype=str)
    df.columns = df.columns.str.strip()
    df = df[df["Beneficio"].notna() & (df["Beneficio"].str.strip() != "")]
    return df.reset_index(drop=True)

def login(page):
    print("[1/3] Abriendo PAMI...")
    page.goto("https://efectoresweb.pami.org.ar/EfectoresWeb/login.isp")
    page.wait_for_selector('input[type="text"]', timeout=15000)

    print("[2/3] Iniciando sesión...")
    page.locator('input[type="text"]').first.fill(USUARIO)
    page.locator('input[type="password"]').first.fill(CLAVE)
    page.get_by_role("button", name="INICIAR SESION").click()

    try:
        page.wait_for_selector("text=Usuario y/o contraseña incorrecta.", timeout=4000)
        page.locator(".z-messagebox-button").click()  # cerrar el dialog de error
        raise LoginError("Usuario y/o contraseña incorrecta. Verificá las credenciales.")
    except PWTimeout:
        pass  # no apareció el error → login exitoso

    print("[2/3] Login exitoso.")

def ir_a_ambulatorio(page):
    print("[3/3] Navegando a Prestaciones Ambulatorias...")
    page.goto("https://efectoresweb.pami.org.ar/EfectoresWeb/ambulatorio.isp")
    page.wait_for_selector("text=ALTA", timeout=15000)
    pausa()

# ── Afiliado ──────────────────────────────────────────────────────────────────

def cargar_afiliado(page, beneficio, parentesco):
    print(f"  Afiliado: {beneficio} | parentesco: {parentesco}")

    page.locator("#zk_comp_130-btn").click()
    popup = page.locator("#zk_comp_130-pp")
    try:
        popup.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise OrdenError("El panel de búsqueda de afiliado no se abrió. La página puede estar lenta o en un estado inesperado.")
    pausa()

    page.locator("#zk_comp_153").click()
    page.locator("#zk_comp_153").press("Control+a")
    page.locator("#zk_comp_153").press("Delete")
    page.locator("#zk_comp_153").type(beneficio, delay=TYPING_MS)
    pausa_corta()

    cod_parentesco = parentesco.split(" - ")[0] if " - " in parentesco else parentesco
    cod_parentesco = cod_parentesco.strip().zfill(2)  # garantiza 2 dígitos: "1" → "01"

    # Esperamos el botón flecha del combobox (abre el dropdown); el input solo pone foco
    parentesco_btn = popup.locator("tr").filter(has_text="Parentesco").locator(".z-combobox-button")
    try:
        parentesco_btn.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise OrdenError(f"El selector de parentesco no apareció (beneficio: '{beneficio}'). La página puede estar lenta.")
    parentesco_btn.click()
    pausa_corta()

    dropdown = page.locator(".z-combobox-popup.z-combobox-open")
    try:
        dropdown.wait_for(state="visible", timeout=5000)
    except PWTimeout:
        raise OrdenError(f"El dropdown de parentesco no se abrió (beneficio: '{beneficio}').")

    items_parentesco = dropdown.locator(".z-comboitem-text", has_text=re.compile(rf"^{cod_parentesco}"))
    if items_parentesco.count() == 0:
        page.keyboard.press("Escape")
        pausa_corta()
        raise OrdenError(f"Código de parentesco '{cod_parentesco}' no encontrado en PAMI (beneficio: '{beneficio}').")
    items_parentesco.first.click()
    pausa()

    popup.get_by_role("button", name="Buscar").click()

    primer_item = popup.locator(".z-listitem").first
    try:
        primer_item.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise BeneficioNoEncontrado(f"Beneficio '{beneficio}' con parentesco '{parentesco}' no encontrado en PAMI.")

    primer_item.click()
    pausa()

# ── Fecha ─────────────────────────────────────────────────────────────────────

def cargar_fecha(page, fecha_str):
    print(f"  Fecha: {fecha_str}")
    fecha = pd.to_datetime(fecha_str.strip(), dayfirst=True).to_pydatetime()

    page.locator("#zk_comp_128-real").click()
    popup = page.locator("#zk_comp_128-pp")
    try:
        popup.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise OrdenError(f"El calendario de fechas no se abrió al cargar la fecha '{fecha_str}'.")
    time.sleep(0.3)  # esperar render inicial del calendario — fijo, independiente del perfil

    # El calendario abre siempre en el mes actual — calculamos cuántos meses navegar
    hoy  = date.today()
    diff = (fecha.year - hoy.year) * 12 + (fecha.month - hoy.month)

    if diff < 0:
        for _ in range(abs(diff)):
            popup.locator("[id$='-left']").click()
            time.sleep(0.55)  # ZK anima la transición de mes (~300-500ms) — fijo para garantizar estabilidad
    elif diff > 0:
        for _ in range(diff):
            popup.locator("[id$='-right']").click()
            time.sleep(0.55)

    # Clickeamos el día exacto, excluyendo días del mes adyacente.
    # Se usa .first como seguro: si la animación aún no terminó y hay 2 grillas en el DOM,
    # no se lanza strict mode violation. Con 0.55s de espera el caso normal tiene 1 sola celda.
    popup.locator("td.z-calendar-cell:not(.z-calendar-outside)").filter(
        has_text=re.compile(rf"^{fecha.day}$")
    ).first.click()
    pausa()

    try:
        page.wait_for_selector("text=La fecha/hora de la prestación no puede superar", timeout=2000)
        page.locator(".z-messagebox-button").click()
        pausa_corta()
        raise OrdenError(f"La fecha {fecha_str} supera la fecha actual. PAMI no permite cargar fechas futuras.")
    except PWTimeout:
        pass  # fecha válida

# ── Profesional Actuante ──────────────────────────────────────────────────────

def cargar_profesional(page):
    print("  Profesional actuante...")
    page.locator("#zk_comp_380-btn").click()
    try:
        page.locator("#zk_comp_382").wait_for(state="visible", timeout=8000)
    except PWTimeout:
        raise OrdenError("El panel de selección de profesional no respondió.")
    pausa_corta()
    page.locator("#zk_comp_382").click()
    pausa_corta()

# ── Diagnóstico ───────────────────────────────────────────────────────────────

def cargar_diagnostico(page, cod_diagnostico):
    print(f"  Diagnóstico: {cod_diagnostico}")

    page.locator("#zk_comp_223-real").click()
    popup = page.locator("#zk_comp_223-pp")
    try:
        popup.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise OrdenError(f"El panel de búsqueda de diagnóstico no se abrió al cargar '{cod_diagnostico}'.")
    pausa()

    popup.locator("#zk_comp_236").click()
    popup.locator("#zk_comp_236").press("Control+a")
    popup.locator("#zk_comp_236").press("Delete")
    popup.locator("#zk_comp_236").type(cod_diagnostico, delay=TYPING_MS)
    pausa_corta()

    popup.get_by_role("button", name="Buscar").click()
    try:
        popup.locator(".z-listitem").first.wait_for(state="visible", timeout=12000)
    except PWTimeout:
        page.keyboard.press("Escape")
        pausa_corta()
        raise DiagnosticoNoEncontrado(f"Código de diagnóstico '{cod_diagnostico.upper()}' no encontrado.")
    pausa_corta()

    items = popup.locator(".z-listitem")
    primer_cod = items.first.locator(".z-listcell").first.inner_text().strip() if items.count() > 0 else ""
    if primer_cod != cod_diagnostico.upper():
        page.keyboard.press("Escape")
        pausa_corta()
        raise DiagnosticoNoEncontrado(f"Código de diagnóstico '{cod_diagnostico.upper()}' no encontrado.")

    items.first.click()
    pausa()

    # PRIMARIO ya viene seleccionado por defecto; click en AGREGAR
    page.locator("#zk_comp_262").click()
    pausa()

# ── Práctica ──────────────────────────────────────────────────────────────────

def cargar_practica(page, cod_practica):
    print(f"  Práctica: {cod_practica}")

    # Abrir el bandbox de prácticas
    page.locator("#zk_comp_280-real").click()
    popup = page.locator("#zk_comp_280-pp")
    try:
        popup.wait_for(state="visible", timeout=10000)
    except PWTimeout:
        raise OrdenError(f"El panel de búsqueda de prácticas no se abrió al cargar '{cod_practica}'.")
    pausa_corta()

    # Escribir el código simulando tipeo humano
    campo = popup.locator("#zk_comp_285")
    campo.click()
    campo.press("Control+a")
    campo.press("Delete")
    campo.type(cod_practica, delay=TYPING_MS)
    campo.press("Enter")
    try:
        popup.locator(".z-listitem").first.wait_for(state="visible", timeout=12000)
    except PWTimeout:
        page.keyboard.press("Escape")
        pausa_corta()
        raise PracticaNoEncontrada(f"Código de práctica '{cod_practica}' no encontrado.")
    pausa_corta()

    # Verificar que el primer resultado sea coincidencia exacta (el buscador hace match parcial)
    items = popup.locator(".z-listitem")
    primer_cod = items.first.locator(".z-listcell").first.inner_text().strip() if items.count() > 0 else ""
    if primer_cod != cod_practica:
        page.keyboard.press("Escape")
        pausa_corta()
        raise PracticaNoEncontrada(f"Código de práctica '{cod_practica}' no encontrado.")

    items.first.click()
    pausa_corta()

    # Hora: forzamos 00:00 seleccionando todo y escribiendo
    hora = page.locator("#zk_comp_303-real")
    hora.click()
    hora.press("Home")
    page.keyboard.type("00")       # horas
    hora.press("ArrowRight")
    page.keyboard.type("00")       # minutos
    pausa_corta()

    # Cantidad: 1
    page.locator("#zk_comp_306").click()
    page.locator("#zk_comp_306").fill("1")
    pausa_corta()

    # Modalidad: Afiliado Propio — selección por texto para no depender del ID del item
    page.locator("#zk_comp_308-real").click()
    pausa_corta()
    page.locator("#zk_comp_308-btn").click()
    pausa_corta()
    modal_dropdown = page.locator(".z-combobox-popup.z-combobox-open")
    try:
        modal_dropdown.wait_for(state="visible", timeout=5000)
    except PWTimeout:
        raise OrdenError(f"El dropdown de modalidad no se abrió para práctica '{cod_practica}'.")
    modal_dropdown.locator(".z-comboitem-text", has_text="AFILIADO PROPIO").click()
    pausa_corta()

    # Agregar
    page.locator("#zk_comp_313").click()
    pausa()

    # Cerrar popup de validación si aparece (ej: "Debe seleccionar una modalidad")
    err_popup = page.locator(".z-messagebox")
    if err_popup.is_visible():
        mensaje = err_popup.locator(".z-messagebox-cnt").inner_text().strip()
        page.locator(".z-messagebox-button").click()
        pausa_corta()
        raise OrdenError(f"Error al agregar práctica '{cod_practica}': {mensaje}")

# ── Orden completa ────────────────────────────────────────────────────────────

def cancelar_orden(page):
    page.locator("#zk_comp_318").click()  # CANCELAR_ORDEN
    page.wait_for_selector("text=ALTA", state="visible", timeout=15000)

def nueva_orden(page, fila):
    print(f"  Iniciando ALTA para beneficio {fila['Beneficio']}...")

    page.locator("text=ALTA").first.click()
    try:
        page.wait_for_selector("#zk_comp_130-btn", state="visible", timeout=20000)
    except PWTimeout:
        raise OrdenError("El formulario de alta no se cargó tras hacer clic en ALTA. La página puede estar lenta.")
    pausa()

    try:
        cargar_afiliado(page, fila["Beneficio"], fila["Parentesco"])
        check_stop()
        cargar_fecha(page, fila["Fecha"])
        check_stop()
        cargar_profesional(page)
        check_stop()
        cargar_diagnostico(page, fila["Cod_Diagnostico"])
        check_stop()
        practica_cols = sorted(c for c in fila.index if c.startswith("Cod_Practica"))
        practicas = [str(fila[col]).strip() for col in practica_cols
                     if pd.notna(fila[col]) and str(fila[col]).strip()]
        if not practicas:
            raise PracticaNoEncontrada("No hay códigos de práctica válidos para esta fila.")
        for i, cod in enumerate(practicas):
            if i > 0:
                page.locator("#zk_comp_279-cave").click()
                pausa_corta()
            cargar_practica(page, cod)
            check_stop()
    except Exception:
        try:
            cancelar_orden(page)
        except Exception as e_cancel:
            print(f"  [AVISO] No se pudo cancelar la orden: {e_cancel}. Intentando recuperar...")
            try:
                ir_a_ambulatorio(page)
            except Exception:
                pass
        raise

    if DRY_RUN:
        print("  [MODO PRUEBA] Cancelando sin guardar...")
        cancelar_orden(page)
        print(f"  -> PRUEBA: beneficio {fila['Beneficio']}")
    else:
        print("  Guardando orden...")
        page.locator("#zk_comp_317").click()  # ACEPTAR_ORDEN
        loc_alta      = page.locator("text=ALTA")
        loc_duplicado = page.locator("text=No puede existir mas de un ambulatorio")
        try:
            loc_alta.or_(loc_duplicado).first.wait_for(state="visible", timeout=15000)
        except PWTimeout:
            raise OrdenError("La orden fue enviada pero PAMI no confirmó el resultado. Verificá manualmente si quedó registrada.")
        if loc_duplicado.is_visible():
            page.locator(".z-messagebox-button").click()  # cerrar popup duplicado
            pausa_corta()
            cancelar_orden(page)
            raise SesionDuplicada("Ya existe una orden para este afiliado/profesional en esa fecha.")
        print(f"  -> OK: beneficio {fila['Beneficio']}")

# ── Main ──────────────────────────────────────────────────────────────────────

def run(playwright: Playwright) -> None:
    df = leer_pacientes()

    browser = playwright.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO)
    context = browser.new_context()
    page    = context.new_page()

    try:
        try:
            login(page)
        except LoginError as e:
            print(f"\n[ERROR FATAL] {e}")
            print("El bot se detuvo. Corregí las credenciales y volvé a intentarlo.")
            return

        ir_a_ambulatorio(page)

        resultados = []
        detenido   = False

        for idx, fila in df.iterrows():
            if STOP_FLAG.exists():
                STOP_FLAG.unlink(missing_ok=True)
                detenido = True
                print("[DETENIDO] Ejecución detenida por el usuario.")
                for fila_pend in df.iloc[idx:].itertuples():
                    resultados.append({
                        "beneficio": fila_pend.Beneficio,
                        "estado":    "DETENIDO",
                        "motivo":    "Detenido por el usuario",
                        "_df_idx":   fila_pend.Index,
                    })
                break

            beneficio = fila["Beneficio"]
            print(f"\n=== Fila {idx + 1} de {len(df)} | Beneficio {beneficio} ===")

            try:
                fecha_fila = pd.to_datetime(fila["Fecha"].strip(), dayfirst=True).date()
                if fecha_fila > date.today():
                    print(f"  [OMITIDO] Fecha futura: {fila['Fecha']}")
                    resultados.append({"beneficio": beneficio, "estado": "OMITIDO", "motivo": f"Fecha futura: {fila['Fecha']}", "_df_idx": idx})
                    continue
            except Exception:
                pass

            try:
                nueva_orden(page, fila)
                estado = "PRUEBA" if DRY_RUN else "OK"
                resultados.append({"beneficio": beneficio, "estado": estado, "motivo": "", "_df_idx": idx})
            except DetenerError:
                print(f"  [DETENIDO] Parada solicitada — orden {beneficio} cancelada.")
                resultados.append({"beneficio": beneficio, "estado": "DETENIDO", "motivo": "Detenido por el usuario", "_df_idx": idx})
                for fila_pend in df.iloc[idx + 1:].itertuples():
                    resultados.append({
                        "beneficio": fila_pend.Beneficio,
                        "estado":    "DETENIDO",
                        "motivo":    "Detenido por el usuario",
                        "_df_idx":   fila_pend.Index,
                    })
                detenido = True
                STOP_FLAG.unlink(missing_ok=True)
                break
            except OrdenError as e:
                print(f"  [FALLO] {e}")
                resultados.append({"beneficio": beneficio, "estado": "ERROR", "motivo": str(e), "_df_idx": idx})
            except Exception as e:
                print(f"  [ERROR INESPERADO] {e}")
                resultados.append({"beneficio": beneficio, "estado": "ERROR", "motivo": f"Error inesperado: {e}", "_df_idx": idx})

        if not detenido and RETRIES > 0:
            a_reintentar = [
                (i, df.loc[r["_df_idx"]])
                for i, r in enumerate(resultados)
                if r["estado"] == "ERROR"
            ]
            if a_reintentar:
                linea_sep = "─" * 50
                print(f"\n{linea_sep}")
                print(f"REINTENTANDO {len(a_reintentar)} fila(s) con error...")
                print(linea_sep)
                try:
                    ir_a_ambulatorio(page)
                except Exception as e:
                    print(f"  [AVISO] No se pudo navegar a ambulatorio antes del reintento: {e}. Se omiten reintentos.")
                    a_reintentar = []
                for res_idx, fila_r in a_reintentar:
                    if STOP_FLAG.exists():
                        STOP_FLAG.unlink(missing_ok=True)
                        break
                    beneficio_r = fila_r["Beneficio"]
                    print(f"\n=== [REINTENTO] Beneficio {beneficio_r} ===")
                    try:
                        nueva_orden(page, fila_r)
                        resultados[res_idx]["estado"] = "PRUEBA" if DRY_RUN else "OK"
                        resultados[res_idx]["motivo"] = "Recuperado en reintento"
                    except SesionDuplicada:
                        resultados[res_idx]["estado"] = "PRUEBA" if DRY_RUN else "OK"
                        resultados[res_idx]["motivo"] = "Primer intento guardado (detectado por duplicado en reintento)"
                    except LoginError as e:
                        print(f"  [SESIÓN EXPIRADA] {e} — se interrumpen los reintentos.")
                        break
                    except DetenerError:
                        STOP_FLAG.unlink(missing_ok=True)
                        break
                    except Exception as e:
                        resultados[res_idx]["motivo"] += f" | Reintento: {e}"
                        print(f"  [REINTENTO FALLIDO] {e}")

        if detenido:
            print("REPORTE PARCIAL — procesamiento detenido antes de completar.")

        sep  = "=" * 50
        ok   = sum(1 for r in resultados if r["estado"] in ("OK", "PRUEBA"))
        omit = sum(1 for r in resultados if r["estado"] == "OMITIDO")
        det  = sum(1 for r in resultados if r["estado"] == "DETENIDO")
        err  = len(resultados) - ok - omit - det
        print(f"\n{sep}")
        print("RESUMEN FINAL")
        print(sep)
        print(f"Total: {len(resultados)} | OK: {ok} | Omitidos: {omit} | Detenidos: {det} | Errores: {err}")
        if omit:
            print(f"\n{omit} fila(s) omitidas por fecha futura (procesables cuando llegue su fecha).")
        if err:
            print("\nDetalle de errores:")
            for r in resultados:
                if r["estado"] == "ERROR":
                    print(f"  - Beneficio {r['beneficio']}: {r['motivo']}")
        print(sep)

        # ── Guardar reporte Excel con colores ─────────────────────────────────────
        try:
            df["Estado"] = [r["estado"] for r in resultados]
            df["Motivo"] = [r["motivo"] for r in resultados]

            practica_cols = sorted(c for c in df.columns if c.startswith("Cod_Practica"))
            otras_cols    = [c for c in df.columns if c not in practica_cols and c not in ("Estado", "Motivo")]
            df[""]        = ""  # separador visual
            df = df[["Estado", "Motivo", ""] + otras_cols + practica_cols]

            PAMI_REPORTES.mkdir(parents=True, exist_ok=True)
            reporte_path = PAMI_REPORTES / f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            df.to_excel(reporte_path, index=False)

            wb = load_workbook(reporte_path)
            ws = wb.active
            fill_ok   = PatternFill("solid", fgColor="C6EFCE")  # verde suave
            fill_err  = PatternFill("solid", fgColor="FFC7CE")  # rojo suave
            fill_omit = PatternFill("solid", fgColor="FFEB9C")  # amarillo suave
            fill_pend = PatternFill("solid", fgColor="D9D9D9")  # gris suave
            col_estado = next(c.column for c in ws[1] if c.value == "Estado")
            for row in ws.iter_rows(min_row=2, min_col=col_estado, max_col=col_estado):
                for cell in row:
                    if cell.value in ("OK", "PRUEBA"):
                        cell.fill = fill_ok
                    elif cell.value == "OMITIDO":
                        cell.fill = fill_omit
                    elif cell.value in ("PENDIENTE", "DETENIDO"):
                        cell.fill = fill_pend
                    else:
                        cell.fill = fill_err
            wb.save(reporte_path)
            print(f"Reporte guardado en {reporte_path}")
        except PermissionError:
            print("[ERROR REPORTE] No se pudo guardar el reporte: el archivo está abierto en Excel o sin permisos de escritura.")
        except OSError as e:
            print(f"[ERROR REPORTE] No se pudo guardar el reporte (error de sistema): {e}")
        except Exception as e:
            print(f"[ERROR REPORTE] Error inesperado al generar el reporte: {e}")

        print("\nProceso finalizado.")

    finally:
        context.close()
        browser.close()

if __name__ == "__main__":
    with sync_playwright() as playwright:
        run(playwright)
